import csv
import os
import shutil
import subprocess
import sys
import tkinter as tk
from datetime import datetime
from tkinter import ttk, messagebox, filedialog
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.styles import Font, Alignment, PatternFill
import pandas as pd
from PIL import Image, ImageTk
from openpyxl import Workbook
import io


class TradeJournalApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Trade Journal")
        self.filename = "trade_journal.csv"
        self.screenshot_dir = "trade_screenshots"
        self.current_screenshots = []

        # Create screenshot directory if it doesn't exist
        os.makedirs(self.screenshot_dir, exist_ok=True)

        self._setup_ui()
        self._initialize_file()

    def _initialize_file(self):
        """Create CSV file with headers if it doesn't exist"""
        if not os.path.exists(self.filename):
            with open(self.filename, 'w', newline='') as f:
                writer = csv.writer(f)
                writer.writerow([
                    'Timestamp', 'Instrument', 'Direction', 'Entry', 'Exit',
                    'Stop Loss', 'Take Profit', 'Size', 'Risk', 'Reward',
                    'P/L', 'Duration', 'Strategy', 'Setup', 'Mistakes',
                    'Lessons', 'Screenshots'
                ])

    def _setup_ui(self):
        """Create the user interface with screenshot support"""
        # Main frame
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))

        # Form fields
        fields = [
            ('Instrument', 'NQ'),
            ('Direction', ['Long', 'Short']),
            ('Entry', '18050.25'),
            ('Exit', '18100.50'),
            ('Stop Loss', '18025.00'),
            ('Take Profit', '18150.00'),
            ('Size', '1'),
            ('Risk', '50'),
            ('Reward', '100'),
            ('P/L', '50.25'),
            ('Duration', '2h15m'),
            ('Strategy', 'EMA Pullback'),
            ('Setup', ''),
            ('Mistakes', ''),
            ('Lessons', '')
        ]

        self.entries = {}

        for i, (label, default) in enumerate(fields):
            ttk.Label(main_frame, text=label).grid(row=i, column=0, sticky=tk.W)

            if isinstance(default, list):  # Dropdown for Direction
                self.entries[label] = ttk.Combobox(main_frame, values=default)
                self.entries[label].set(default[0])
            else:
                self.entries[label] = ttk.Entry(main_frame)
                self.entries[label].insert(0, default)

            self.entries[label].grid(row=i, column=1, sticky=(tk.W, tk.E))

        # Screenshot section
        ttk.Label(main_frame, text="Screenshots").grid(row=len(fields), column=0, sticky=tk.W)

        self.screenshot_frame = ttk.Frame(main_frame)
        self.screenshot_frame.grid(row=len(fields), column=1, sticky=(tk.W, tk.E))

        add_screenshot_btn = ttk.Button(
            main_frame,
            text="Add Screenshot",
            command=self._add_screenshot
        )
        add_screenshot_btn.grid(row=len(fields) + 1, column=0, columnspan=2, pady=5)

        # Submit button
        submit_btn = ttk.Button(main_frame, text="Log Trade", command=self._submit)
        submit_btn.grid(row=len(fields) + 2, column=0, columnspan=2, pady=10)

        export_frame = ttk.Frame(self.root, padding="10")
        export_frame.grid(row=1, column=0, sticky=(tk.W, tk.E))

        # Add export button
        ttk.Button(
            export_frame,
            text="Export to Excel",
            command=self.export_to_excel,
            style='Accent.TButton'
        ).pack(side=tk.LEFT, padx=5)

        ttk.Button(
            export_frame,
            text="Open Excel File",
            command=self.open_excel_file
        ).pack(side=tk.LEFT, padx=5)

        # Configure style for accent button
        style = ttk.Style()
        style.configure('Accent.TButton', foreground='white', background='#4F81BD')
        # ok?


        # Configure grid weights
        main_frame.columnconfigure(1, weight=1)

    def _add_screenshot(self):
        """Open file dialog to add screenshot"""
        filetypes = (
            ('Image files', '*.png *.jpg *.jpeg'),
            ('All files', '*.*')
        )

        filenames = filedialog.askopenfilenames(
            title="Select Screenshot(s)",
            initialdir="/",
            filetypes=filetypes
        )

        if filenames:
            for filename in filenames:
                if filename not in self.current_screenshots:
                    self.current_screenshots.append(filename)
                    self._display_screenshot_thumbnail(filename)

    def _display_screenshot_thumbnail(self, filename):
        """Show a thumbnail of the screenshot in the UI"""
        try:
            # Create thumbnail
            img = Image.open(filename)
            img.thumbnail((100, 100))
            photo = ImageTk.PhotoImage(img)

            # Create thumbnail label
            thumb_frame = ttk.Frame(self.screenshot_frame)
            thumb_frame.pack(side=tk.LEFT, padx=5)

            label = ttk.Label(thumb_frame, image=photo)
            label.image = photo  # Keep reference
            label.pack()

            # Add filename label
            filename_short = os.path.basename(filename)[:15] + "..."
            ttk.Label(thumb_frame, text=filename_short).pack()

            # Add remove button
            remove_btn = ttk.Button(
                thumb_frame,
                text="X",
                command=lambda f=filename, tf=thumb_frame: self._remove_screenshot(f, tf)
            )
            remove_btn.pack()

        except Exception as e:
            messagebox.showerror("Error", f"Could not load image: {str(e)}")

    def _remove_screenshot(self, filename, thumb_frame):
        """Remove a screenshot from the current selection"""
        if filename in self.current_screenshots:
            self.current_screenshots.remove(filename)
        thumb_frame.destroy()

    def _submit(self):
        """Handle form submission with screenshots"""
        try:
            # Save screenshots to dedicated folder
            saved_screenshot_paths = []
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')

            for i, screenshot_path in enumerate(self.current_screenshots):
                if os.path.exists(screenshot_path):
                    # Create unique filename
                    ext = os.path.splitext(screenshot_path)[1]
                    new_filename = f"{timestamp}_{i}{ext}"
                    dest_path = os.path.join(self.screenshot_dir, new_filename)

                    # Copy the file
                    shutil.copy2(screenshot_path, dest_path)
                    saved_screenshot_paths.append(dest_path)

            # Prepare trade data
            trade_data = {
                'Timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'Instrument': self.entries['Instrument'].get(),
                'Direction': self.entries['Direction'].get(),
                'Entry': float(self.entries['Entry'].get()),
                'Exit': float(self.entries['Exit'].get()),
                'Stop Loss': float(self.entries['Stop Loss'].get()),
                'Take Profit': float(self.entries['Take Profit'].get()),
                'Size': float(self.entries['Size'].get()),
                'Risk': float(self.entries['Risk'].get()),
                'Reward': float(self.entries['Reward'].get()),
                'P/L': float(self.entries['P/L'].get()),
                'Duration': self.entries['Duration'].get(),
                'Strategy': self.entries['Strategy'].get(),
                'Setup': self.entries['Setup'].get(),
                'Mistakes': self.entries['Mistakes'].get(),
                'Lessons': self.entries['Lessons'].get(),
                'Screenshots': ';'.join(saved_screenshot_paths)  # Store paths as semicolon-separated string
            }

            # Write to CSV
            with open(self.filename, 'a', newline='') as f:
                writer = csv.DictWriter(f, fieldnames=trade_data.keys())
                writer.writerow(trade_data)

            messagebox.showinfo("Success", "Trade logged successfully!")

            # Clear form
            for entry in self.entries.values():
                if isinstance(entry, ttk.Combobox):
                    entry.set(entry['values'][0])
                else:
                    entry.delete(0, tk.END)

            # Clear screenshots
            self.current_screenshots = []
            for widget in self.screenshot_frame.winfo_children():
                widget.destroy()

        except ValueError as e:
            messagebox.showerror("Error", f"Invalid input: {str(e)}")
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred: {str(e)}")


    # export functionh
    def export_to_excel(self):
        """Export CSV data to Excel with screenshots"""
        try:
            output_file = "trades_export.xlsx"

            if not os.path.exists(self.filename):
                messagebox.showerror("Error", "No trade journal CSV file found!")
                return

            # Load CSV
            df = pd.read_csv(self.filename)

            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Trade Journal"

            headers = list(df.columns) + ["Screenshot Preview"]
            ws.append(headers)

            # Format headers
            header_style = {
                'fill': PatternFill(start_color='4F81BD', fill_type='solid'),
                'font': Font(color='FFFFFF', bold=True),
                'alignment': Alignment(horizontal='center')
            }

            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=1, column=col)
                for attr, value in header_style.items():
                    setattr(cell, attr, value)

            # Process trades
            for idx, row in df.iterrows():
                try:
                    # Write trade data
                    ws.append(row.tolist())
                    current_row = idx + 2  # +1 for header, +1 for 0-index

                    # Handle screenshots
                    screenshot_paths = []
                    if pd.notna(row['Screenshots']):
                        screenshot_paths = [p.strip() for p in str(row['Screenshots']).split(';') if p.strip()]

                    if screenshot_paths:
                        # Process first screenshot
                        img_path = screenshot_paths[0]
                        try:
                            # Validate and load image
                            if not os.path.exists(img_path):
                                print(f"Image not found: {img_path}")
                                continue

                            # Load and resize image
                            with Image.open(img_path) as pil_img:
                                pil_img.thumbnail((300, 300))

                                # Convert to bytes
                                img_bytes = io.BytesIO()
                                pil_img.save(img_bytes, format='PNG')
                                img_bytes.seek(0)

                                # Add to worksheet
                                img = OpenpyxlImage(img_bytes)
                                ws.add_image(img, f"Q{current_row}")  # Column Q

                                # Note additional screenshots
                                if len(screenshot_paths) > 1:
                                    note_cell = ws.cell(row=current_row, column=len(headers))
                                    note_cell.value = f"+{len(screenshot_paths) - 1} more"

                        except Exception as img_error:
                            print(f"Error processing image {img_path}: {str(img_error)}")

                except Exception as row_error:
                    print(f"Error processing row {idx}: {str(row_error)}")
                    continue

            # Adjust layout
            try:
                # Column widths
                for col in ws.columns:
                    max_len = max(
                        (len(str(cell.value)) for cell in col),
                        default=10
                    )
                    ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

                # Row heights
                for row in range(2, len(df) + 2):
                    ws.row_dimensions[row].height = 120

                # Freeze headers
                ws.freeze_panes = "A2"

                # Save
                wb.save(output_excel)
                print(f"Successfully saved to {output_excel}")

            except Exception as save_error:
                print(f"Error saving Excel file: {str(save_error)}")
            # Save workbook
            wb.save(output_file)
            messagebox.showinfo("Success", f"Exported to {output_file}")
            return output_file

        except Exception as e:
            messagebox.showerror("Error", f"Export failed:\n{str(e)}")
            return None

    def open_excel_file(self):
        """Open the exported Excel file"""
        try:
            output_file = "trades_export.xlsx"

            if not os.path.exists(output_file):
                # Auto-export if file doesn't exist
                output_file = self.export_to_excel()
                if not output_file:
                    return

            # Platform-independent file opening
            if sys.platform == "win32":
                os.startfile(output_file)
            elif sys.platform == "darwin":
                subprocess.run(["open", output_file])
            else:
                subprocess.run(["xdg-open", output_file])

        except Exception as e:
            messagebox.showerror("Error", f"Could not open file:\n{str(e)}")
# Run the application
if __name__ == "__main__":
    root = tk.Tk()
    app = TradeJournalApp(root)
    root.mainloop()
